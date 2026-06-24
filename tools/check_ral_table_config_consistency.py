from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"D:\home\RAL_paper\RAL_data_result\outputs\RAL_experiment_results_summary_academic.xlsx")
ROOT = Path(r"D:\home\catkin_ws-main\paper_experiments")


def read_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    pattern = re.compile(r'^([A-Z0-9_]+)="?\$\{[^:}]+:-(.*?)\}"?$')
    direct = re.compile(r'^([A-Z0-9_]+)="?([^"#]+?)"?$')
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        match = pattern.match(line)
        if match:
            values[match.group(1)] = match.group(2)
            continue
        match = direct.match(line)
        if match:
            values[match.group(1)] = match.group(2)
    return values


def read_orb_yaml(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or line.startswith("%"):
            continue
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        value = value.strip().strip('"')
        values[key.strip()] = value
    return values


def compact(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def dump_config_sheets():
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    sheets = {}
    for ws in wb.worksheets:
        if not ws.title.lower().endswith("_config"):
            continue
        rows = []
        for row in ws.iter_rows(values_only=True):
            row = [compact(v) for v in row]
            if any(v is not None for v in row):
                rows.append(row)
        sheets[ws.title] = rows
    return sheets


def main() -> None:
    configs = {
        "exp3": {
            "env": read_env(ROOT / "exp3_aquaticvision_generalization/run_config/exp3_config.env"),
            "yaml": {
                "aquaticvision_mono": read_orb_yaml(ROOT / "exp3_aquaticvision_generalization/run_config/aquaticvision_mono.yaml"),
                "aquaticvision_stereo": read_orb_yaml(ROOT / "exp3_aquaticvision_generalization/run_config/aquaticvision_stereo.yaml"),
            },
        },
        "exp4": {
            "env": read_env(ROOT / "exp4_euroc_sota_comparison/run_config/exp4_config.env"),
            "yaml": {
                "euroc_mono_inertial": read_orb_yaml(ROOT / "exp4_euroc_sota_comparison/run_config/euroc_mono_inertial.yaml"),
                "euroc_stereo_inertial": read_orb_yaml(ROOT / "exp4_euroc_sota_comparison/run_config/euroc_stereo_inertial.yaml"),
            },
        },
        "exp5": {
            "env": read_env(ROOT / "exp5_efficiency_resource/run_config/exp5_config.env"),
            "yaml": {
                p.stem: read_orb_yaml(p)
                for p in sorted((ROOT / "exp5_efficiency_resource/run_config").glob("*.yaml"))
            },
        },
    }

    print("=== Workbook config sheets ===")
    print(json.dumps(dump_config_sheets(), ensure_ascii=False, indent=2))
    print("=== Project run_config values ===")
    print(json.dumps(configs, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
